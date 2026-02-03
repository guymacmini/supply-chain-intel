"""Excel/CSV export utilities for research data."""

import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


class ExcelExporter:
    """Export research data to Excel and CSV formats."""
    
    def __init__(self, output_dir: Path = None):
        """Initialize exporter.
        
        Args:
            output_dir: Directory for export files. Defaults to current working directory.
        """
        self.output_dir = output_dir or Path.cwd()
        self.output_dir.mkdir(exist_ok=True)
    
    def extract_company_data_from_research(self, research_content: str) -> List[Dict[str, Any]]:
        """Extract structured company/ticker data from research markdown.
        
        Args:
            research_content: Full research document content
            
        Returns:
            List of dictionaries with company data
        """
        companies = []
        
        # Extract markdown tables that contain ticker information
        table_pattern = r'\|([^|]+\|)+[^|]+\|'
        table_lines = []
        
        lines = research_content.split('\n')
        in_table = False
        current_table_lines = []
        
        for line in lines:
            if '|' in line and '-----' in line:
                # Table header separator - start of table
                in_table = True
                # Add the previous line as header if it contains pipes
                if current_table_lines and '|' in current_table_lines[-1]:
                    table_lines.append(current_table_lines[-1])
                table_lines.append(line)
                current_table_lines = []
            elif in_table and '|' in line:
                table_lines.append(line)
            elif in_table and '|' not in line:
                # End of table
                in_table = False
                if table_lines:
                    companies.extend(self._parse_table(table_lines))
                    table_lines = []
            else:
                current_table_lines.append(line)
        
        # Handle table at end of document
        if table_lines:
            companies.extend(self._parse_table(table_lines))
        
        # Also extract tickers mentioned in text
        companies.extend(self._extract_tickers_from_text(research_content))
        
        # Remove duplicates by ticker
        seen_tickers = set()
        unique_companies = []
        for company in companies:
            if company.get('ticker') and company['ticker'] not in seen_tickers:
                seen_tickers.add(company['ticker'])
                unique_companies.append(company)
        
        return unique_companies
    
    def _parse_table(self, table_lines: List[str]) -> List[Dict[str, Any]]:
        """Parse a markdown table into structured data."""
        if len(table_lines) < 3:  # Need header, separator, and at least one row
            return []
        
        # Parse header
        header_line = table_lines[0].strip()
        headers = [h.strip() for h in header_line.split('|') if h.strip()]
        
        # Skip separator line (index 1)
        companies = []
        
        for line in table_lines[2:]:
            if not line.strip() or '-----' in line:
                continue
                
            cells = [cell.strip() for cell in line.split('|') if cell.strip()]
            if len(cells) < len(headers):
                continue
                
            company_data = {}
            for i, header in enumerate(headers):
                if i < len(cells):
                    value = cells[i].strip()
                    
                    # Clean up the value
                    value = re.sub(r'\*\*([^*]+)\*\*', r'\1', value)  # Remove markdown bold
                    value = re.sub(r'^\*|\*$', '', value)  # Remove leading/trailing asterisks
                    
                    # Map common header variations
                    header_clean = header.lower().replace(' ', '_')
                    if header_clean in ['ticker', 'symbol']:
                        company_data['ticker'] = value.upper()
                    elif header_clean in ['company', 'name', 'company_name']:
                        company_data['company_name'] = value
                    elif header_clean in ['market_cap', 'marketcap', 'market cap']:
                        company_data['market_cap'] = value
                    elif header_clean in ['role', 'business', 'description']:
                        company_data['role'] = value
                    elif header_clean in ['exposure_score', 'exposure score', 'score']:
                        company_data['exposure_score'] = value
                    elif header_clean in ['price', 'stock_price']:
                        company_data['price'] = value
                    elif header_clean in ['sector', 'industry']:
                        company_data['sector'] = value
                    else:
                        company_data[header_clean] = value
            
            if company_data.get('ticker'):
                companies.append(company_data)
        
        return companies
    
    def _extract_tickers_from_text(self, content: str) -> List[Dict[str, Any]]:
        """Extract ticker symbols mentioned in text."""
        # Look for patterns like (TICKER), **Ticker**, or Ticker Corp (TICK)
        ticker_patterns = [
            r'\(([A-Z]{1,5})\)',  # (AAPL)
            r'\*\*([A-Z]{2,5})\*\*',  # **AAPL**
            r'([A-Z][A-Z0-9]{1,4})(?:\s+Corp|\s+Inc|\s+Ltd|\s+Technologies|\s+Holdings)',  # Apple Corp
        ]
        
        companies = []
        for pattern in ticker_patterns:
            matches = re.findall(pattern, content)
            for match in matches:
                ticker = match.upper()
                if len(ticker) >= 2 and ticker not in ['THE', 'AND', 'FOR', 'ARE', 'BUT', 'NOT']:
                    companies.append({
                        'ticker': ticker,
                        'company_name': 'N/A',
                        'source': 'text_extraction'
                    })
        
        return companies
    
    def extract_metadata_from_research(self, research_content: str) -> Dict[str, Any]:
        """Extract metadata from research document.
        
        Args:
            research_content: Full research document content
            
        Returns:
            Dictionary with metadata
        """
        metadata = {}
        
        # Extract YAML frontmatter
        if research_content.startswith('---'):
            end_marker = research_content.find('---', 3)
            if end_marker != -1:
                frontmatter = research_content[3:end_marker]
                for line in frontmatter.split('\n'):
                    if ':' in line:
                        key, value = line.split(':', 1)
                        key = key.strip()
                        value = value.strip().strip("'\"")
                        metadata[key] = value
        
        # Extract TLDR
        tldr_match = re.search(r'\*\*TLDR:\*\*\s*([^*]+)', research_content, re.IGNORECASE)
        if tldr_match:
            metadata['tldr'] = tldr_match.group(1).strip()
        
        # Extract theme/query
        theme_match = re.search(r'# Investment Research:\s*([^\n]+)', research_content)
        if theme_match:
            metadata['theme'] = theme_match.group(1).strip()
        
        return metadata
    
    def export_research_to_excel(self, research_content: str, filename: str, 
                                output_filename: str = None) -> Path:
        """Export research document to Excel format.
        
        Args:
            research_content: Full research document content
            filename: Original research filename 
            output_filename: Output Excel filename. Auto-generated if None.
            
        Returns:
            Path to created Excel file
        """
        if not output_filename:
            base_name = Path(filename).stem
            output_filename = f"{base_name}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        output_path = self.output_dir / output_filename
        
        # Extract data
        companies = self.extract_company_data_from_research(research_content)
        metadata = self.extract_metadata_from_research(research_content)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create summary sheet
        ws_summary = wb.active
        ws_summary.title = "Research Summary"
        
        # Add summary data
        summary_data = [
            ["Research Document", filename],
            ["Generated", metadata.get('generated', 'N/A')],
            ["Theme", metadata.get('theme', 'N/A')],
            ["Query", metadata.get('query', 'N/A')],
            ["Depth", metadata.get('depth', 'N/A')],
            ["Total Companies", len(companies)],
            ["TLDR", metadata.get('tldr', 'N/A')],
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Auto-adjust column widths
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 50
        
        # Create companies sheet
        if companies:
            ws_companies = wb.create_sheet("Companies")
            
            # Prepare data for companies sheet
            company_df = pd.DataFrame(companies)
            
            # Add header row
            headers = list(company_df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws_companies.cell(row=1, column=col_idx, value=header.replace('_', ' ').title())
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data rows
            for row_idx, (_, row) in enumerate(company_df.iterrows(), 2):
                for col_idx, value in enumerate(row, 1):
                    ws_companies.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            
            # Auto-adjust column widths
            for col_idx in range(1, len(headers) + 1):
                ws_companies.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def export_research_to_csv(self, research_content: str, filename: str,
                              output_filename: str = None) -> Path:
        """Export research document to CSV format.
        
        Args:
            research_content: Full research document content
            filename: Original research filename
            output_filename: Output CSV filename. Auto-generated if None.
            
        Returns:
            Path to created CSV file
        """
        if not output_filename:
            base_name = Path(filename).stem
            output_filename = f"{base_name}_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        output_path = self.output_dir / output_filename
        
        # Extract companies
        companies = self.extract_company_data_from_research(research_content)
        
        if not companies:
            # Create minimal CSV with just metadata
            metadata = self.extract_metadata_from_research(research_content)
            companies = [{'source_file': filename, **metadata}]
        
        # Write to CSV
        if companies:
            fieldnames = set()
            for company in companies:
                fieldnames.update(company.keys())
            fieldnames = sorted(list(fieldnames))
            
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(companies)
        
        return output_path
    
    def export_watchlist_to_excel(self, watchlist_data: List[Dict], 
                                 output_filename: str = None) -> Path:
        """Export watchlist data to Excel format.
        
        Args:
            watchlist_data: List of watchlist entities
            output_filename: Output filename. Auto-generated if None.
            
        Returns:
            Path to created Excel file
        """
        if not output_filename:
            output_filename = f"watchlist_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        output_path = self.output_dir / output_filename
        
        # Convert list fields to strings for Excel compatibility
        cleaned_data = []
        for item in watchlist_data:
            cleaned_item = item.copy()
            if 'themes' in cleaned_item and isinstance(cleaned_item['themes'], list):
                cleaned_item['themes'] = ', '.join(cleaned_item['themes'])
            cleaned_data.append(cleaned_item)
        
        # Create DataFrame
        df = pd.DataFrame(cleaned_data)
        
        # Create workbook and write data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Watchlist"
        
        # Add data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        return output_path
    
    def export_saved_research_to_excel(self, saved_research_data: List[Dict],
                                     output_filename: str = None) -> Path:
        """Export saved research data to Excel format.
        
        Args:
            saved_research_data: List of saved research items
            output_filename: Output filename. Auto-generated if None.
            
        Returns:
            Path to created Excel file  
        """
        if not output_filename:
            output_filename = f"saved_research_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        output_path = self.output_dir / output_filename
        
        # Convert list fields to strings for Excel compatibility
        cleaned_data = []
        for item in saved_research_data:
            cleaned_item = item.copy()
            if 'tags' in cleaned_item and isinstance(cleaned_item['tags'], list):
                cleaned_item['tags'] = ', '.join(cleaned_item['tags'])
            if 'tickers' in cleaned_item and isinstance(cleaned_item['tickers'], list):
                cleaned_item['tickers'] = ', '.join(cleaned_item['tickers'])
            cleaned_data.append(cleaned_item)
        
        # Create DataFrame
        df = pd.DataFrame(cleaned_data)
        
        # Create workbook and write data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Saved Research"
        
        # Add data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        return output_path